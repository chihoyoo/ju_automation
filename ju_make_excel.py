import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_finance_excel(df_finance: pd.DataFrame, df_final: pd.DataFrame | None = None, drive_files: list | None = None, title: str = "정산 리포트", sheet_name: str = "정산") -> tuple[bytes, str]:
    """df_finance를 받아 정산 리포트 형태의 xlsx 바이너리를 반환합니다.

    배치 규칙:
    - B2:H2 타이틀(25pt)
    - B3 브랜드 라벨(10pt)
    - B4:H4 하단 굵은 라인
    - B6:H6 '정산내역' (흰 글씨, 회색 배경)
    - 헤더는 B7부터, 본문은 B8부터
    - 마지막에 소계/계, 그리고 3줄 아래 '실 정산액' 박스
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 스타일
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    title_font = Font(size=16, bold=True)
    bold_font = Font(bold=True)

    # 컬럼 정의 (보이는 순서 고정)
    columns = [
        "상품명",
        "옵션",
        "수량",
        "공구판매가",
        "공구판매가합계(vat포함)",
        "공급가(vat포함)",
        "정산금액(vat포함)",
    ]

    # 배치 시작 위치
    start_col = 2  # B
    end_col = start_col + len(columns) - 1  # H

    # 타이틀: B2:H2
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    tcell = ws.cell(row=2, column=start_col, value=title)
    tcell.font = Font(size=25, bold=True)
    tcell.alignment = center

    # 브랜드 라벨: B3
    ws.cell(row=3, column=start_col, value="소셜라운지").font = Font(size=10)

    # Divider: B4:H4 하단 굵은 선
    thick = Side(border_style="medium", color="000000")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=4, column=c, value=None)
        cell.border = Border(bottom=thick)

    # 정산내역: B6:H6
    ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col)
    s = ws.cell(row=6, column=start_col, value="정산내역")
    s.font = Font(size=10, color="FFFFFF", bold=True)
    s.fill = PatternFill("solid", fgColor="808080")
    s.alignment = center

    # 헤더: B7~H7 (옅은 회색)
    header_row = 7
    for idx, col_name in enumerate(columns):
        c = ws.cell(row=header_row, column=start_col + idx, value=col_name)
        c.fill = PatternFill("solid", fgColor="EDEDED")
        c.font = bold_font
        c.alignment = center
        c.border = border_all

    # 본문: B8부터
    data_row_start = header_row + 1
    current_row = data_row_start
    number_cols = {"수량", "공구판매가", "공구판매가합계(vat포함)", "공급가(vat포함)", "정산금액(vat포함)"}
    df_iter = df_finance if df_finance is not None else pd.DataFrame(columns=columns)
    for _, r in df_iter.iterrows():
        for j, col_name in enumerate(columns):
            val = r.get(col_name)
            c = ws.cell(row=current_row, column=start_col + j, value=val)
            c.border = border_all
            if col_name in number_cols:
                c.number_format = "#,##0"
                c.alignment = right
            else:
                c.alignment = Alignment(vertical="center")
        current_row += 1

    data_row_end = current_row - 1

    # '상품명' 컬럼 병합 (B열): B8 ~ B{data_row_end}
    if data_row_end >= data_row_start:
        ws.merge_cells(start_row=data_row_start, start_column=start_col, end_row=data_row_end, end_column=start_col)
        ws.cell(row=data_row_start, column=start_col).alignment = center

    # 소계 행: 한 줄 추가
    subtotal_row = data_row_end + 1
    if subtotal_row >= data_row_start:
        # 병합 B:C 에 소계
        ws.merge_cells(start_row=subtotal_row, start_column=start_col, end_row=subtotal_row, end_column=start_col + 1)
        sc = ws.cell(row=subtotal_row, column=start_col, value="소계")
        sc.fill = PatternFill("solid", fgColor="EDEDED")
        sc.font = bold_font
        sc.alignment = center
        sc.border = border_all
        # 행 전체 회색 음영 및 테두리
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=subtotal_row, column=c)
            cell.fill = PatternFill("solid", fgColor="EDEDED")
            cell.border = border_all
        # 합계: D, F, H
        col_map = {"수량": 2, "공구판매가합계(vat포함)": 4, "정산금액(vat포함)": 6}  # offset from start_col
        for name, off in col_map.items():
            col_idx = start_col + off
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=subtotal_row, column=col_idx, value=f"=SUM({col_letter}{data_row_start}:{col_letter}{data_row_end})")
            cell.number_format = "#,##0"
            cell.border = border_all
            cell.alignment = right

    # 계 행: 다음 줄
    total_row = subtotal_row + 1
    ws.merge_cells(start_row=total_row, start_column=start_col, end_row=total_row, end_column=end_col - 1)
    tc = ws.cell(row=total_row, column=start_col, value="계")
    tc.fill = PatternFill("solid", fgColor="FFD966")
    tc.font = bold_font
    tc.alignment = center
    # 행 전체 연한 주황색 음영 및 테두리
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = PatternFill("solid", fgColor="FFD966")
        cell.border = border_all
    # H열(정산금액 총계)
    sum_col_idx = start_col + columns.index("정산금액(vat포함)")
    ws.cell(row=total_row, column=end_col, value=f"=SUM({get_column_letter(sum_col_idx)}{data_row_start}:{get_column_letter(sum_col_idx)}{data_row_end})").number_format = "#,##0"

    # 실 정산액 박스: total_row + 3
    final_row = total_row + 3
    # B..C 병합
    ws.merge_cells(start_row=final_row, start_column=start_col, end_row=final_row, end_column=start_col + 1)
    fc = ws.cell(row=final_row, column=start_col, value="실 정산액")
    fc.fill = PatternFill("solid", fgColor="FFC000")
    fc.font = Font(size=12, bold=True)
    fc.alignment = center
    # 굵은 테두리
    thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    for c in range(start_col, start_col + 2):
        ws.cell(row=final_row, column=c).border = thick_border
    # D..H 병합 + 값
    ws.merge_cells(start_row=final_row, start_column=start_col + 2, end_row=final_row, end_column=end_col)
    tv = ws.cell(row=final_row, column=start_col + 2, value=f"=SUM({get_column_letter(sum_col_idx)}{data_row_start}:{get_column_letter(sum_col_idx)}{data_row_end})")
    tv.number_format = "₩#,##0"
    tv.font = Font(bold=True)
    tv.alignment = right
    for c in range(start_col + 2, end_col + 1):
        ws.cell(row=final_row, column=c).border = thick_border

    # 실 정산액 아래 Divider: (final_row + 1) 행의 B..H 아래 테두리 굵게
    divider_row = final_row + 1
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=divider_row, column=c)
        cell.border = Border(bottom=thick)

    # 안내 문구 4줄
    notes_row_start = divider_row + 1
    notes = [
        "*위 금액을 확인하여 주시길 바랍니다.",
        "*이상이 없을 경우 계산서 발행 요청드립니다",
        "사업자 번호 : 790-88-03127",
        "계산서 발행 이메일 : master@sociallounge.company",
    ]
    for i, txt in enumerate(notes):
        ws.cell(row=notes_row_start + i, column=start_col, value=txt)

    # 마지막 Divider 한 줄 더 (안내문구 하단)
    final_divider_row = notes_row_start + len(notes)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=final_divider_row, column=c)
        cell.border = Border(bottom=thick)

    # 범위 외곽 굵은 테두리 (정산내역 ~ 계)
    top = 6
    bottom = total_row
    left = start_col
    right_col = end_col
    for r in range(top, bottom + 1):
        for c in range(left, right_col + 1):
            cell = ws.cell(row=r, column=c)
            new_border = Border(
                left=thick if c == left else cell.border.left,
                right=thick if c == right_col else cell.border.right,
                top=thick if r == top else cell.border.top,
                bottom=thick if r == bottom else cell.border.bottom,
            )
            cell.border = new_border

    # 열 너비 (B..H)
    widths = [22, 38, 10, 14, 20, 16, 18]
    for i, w in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 두번째 시트: raw에 df_final 전체 기록
    if df_final is not None:
        ws2 = wb.create_sheet("raw")
        # 헤더
        for j, col in enumerate(list(df_final.columns), start=1):
            ws2.cell(row=1, column=j, value=str(col))
        # 데이터
        for i, (_, row) in enumerate(df_final.iterrows(), start=2):
            for j, col in enumerate(list(df_final.columns), start=1):
                ws2.cell(row=i, column=j, value=row.get(col))

    # 파일명 구성: 정산서_{yymmdd}_{3}_{4}.xlsx (드라이브 첫 파일명 기준)
    yymmdd = datetime.now().strftime('%y%m%d')
    base_name = ""
    if isinstance(drive_files, list) and drive_files and isinstance(drive_files[0], dict):
        fname = (drive_files[0].get('name') or '').rsplit('.', 1)[0]
        parts = fname.split('_')
        if len(parts) >= 4:
            base_name = f"{parts[2]}_{parts[3]}"
        else:
            base_name = fname
    final_filename = f"정산서_{yymmdd}_{base_name}.xlsx" if base_name else f"정산서_{yymmdd}.xlsx"

    # 저장
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue(), final_filename

